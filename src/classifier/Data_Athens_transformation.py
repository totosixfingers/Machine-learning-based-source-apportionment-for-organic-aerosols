import pandas as pd
import argparse
import os
from openpyxl import load_workbook

def process_excel(input_file, output_dir):
    # Load original sheets
    measure = pd.read_excel(input_file, sheet_name="measurements")
    measure["Date"] = pd.to_datetime(measure["Date"], dayfirst=True)

    # -------------------------
    # Load PMFsolution sheet with openpyxl
    # -------------------------
    wb = load_workbook(input_file)
    ws = wb["PMFsolution"]
    
    header_row = [cell.value for cell in ws[1]]

    # Fill merged header values
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        if min_row == 1:
            top_left_value = ws.cell(row=min_row, column=min_col).value
            for col in range(min_col, max_col + 1):
                header_row[col - 1] = top_left_value
                
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append(row)
        
    pmf = pd.DataFrame(data, columns=header_row)
    # Define periods
    periods = {
    "Warm Period 2016 (26 Jul 2016 - 28 Sep 2016)": ("2016-07-26", "2016-09-28"),
    "Cold period 2016-2017 (01 Nov 2016 - 18 Mar 2017)": ("2016-11-01", "2017-03-18"),
    "Warm Period 2017 (01 May 2017 - 31 Jul 2017)": ("2017-05-01", "2017-07-31")
    }
    os.makedirs(output_dir, exist_ok=True)

    for name, (start, end) in periods.items():
        start = pd.Timestamp(start)
        end = pd.Timestamp(end)
        

        # Filter Measurement for the period
        meas_filtered = measure[(measure["Date"] >= start) & (measure["Date"] <= end)].copy()

        # Create g sheet from PMFsolution filtered by the period
        g_filtered = pmf[(pmf["UTC + 02:00"] >= start) & (pmf["UTC + 02:00"] <= end)].copy()
        g_filtered = g_filtered[["UTC + 02:00", "HOA", "COA", "BBOA", "SV-OOA", "LV-OOA"]]

        # F sheet: same PMF rows but could include extra period columns if needed
        F_filtered = pmf[name].copy()

        # Save Excel with three sheets
        output_file = os.path.join(output_dir, f"{name}.xlsx")
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            meas_filtered.to_excel(writer, sheet_name="measurements", index=False)
            g_filtered.to_excel(writer, sheet_name="G", index=False)
            F_filtered.to_excel(writer, sheet_name="F", index=False, header=False)

        print(f"Saved: {output_file}")


def main():
    parser = argparse.ArgumentParser(description="Split Excel sheets by predefined periods.")
    parser.add_argument("--input", required=True, help="Path to input Excel file.")
    parser.add_argument("--output_dir", required=True, help="Directory to save output files.")
    args = parser.parse_args()


    process_excel(args.input, args.output_dir)


if __name__ == "__main__":
    main()
